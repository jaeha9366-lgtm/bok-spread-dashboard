import os
import json
import csv
import requests
from datetime import datetime, timedelta
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()
api_key = os.getenv("BOK_API_KEY")

def fetch_ecos_data(table_code, item_code, start_date, end_date):
    """
    Fetches daily time-series data from BOK ECOS API.
    Uses index 1 to 10000 to cover 5 years of daily data (approx. 1825 rows max).
    """
    url = f"https://ecos.bok.or.kr/api/StatisticSearch/{api_key}/json/kr/1/10000/{table_code}/D/{start_date}/{end_date}/{item_code}"
    try:
        response = requests.get(url)
        response.raise_for_status()
        data = response.json()
        if "StatisticSearch" in data:
            return data["StatisticSearch"]["row"]
        else:
            print(f"Warning: ECOS API returned no data or error for {table_code} ({item_code}): {data}")
            return []
    except Exception as e:
        print(f"Error fetching data for {table_code} ({item_code}): {e}")
        return []

def get_rate_with_fallback(date_str, data_dict):
    """
    Looks up a date in the data dictionary. If missing (rare holiday mismatch),
    forwards-fills with the most recent preceding value.
    """
    if date_str in data_dict:
        return data_dict[date_str]
    # Fallback: search for the latest available rate before this date
    sorted_dates = sorted([d for d in data_dict.keys() if d < date_str])
    if sorted_dates:
        return data_dict[sorted_dates[-1]]
    return 0.0

def calculate_bond_stats(aligned_data, yield_key, spread_key):
    """
    Calculates detailed statistics for a specific bond yield and spread series.
    """
    spreads = [d[spread_key] for d in aligned_data]
    latest = aligned_data[-1] if aligned_data else {}
    prev_spread = aligned_data[-2][spread_key] if len(aligned_data) > 1 else latest.get(spread_key, 0)
    spread_diff = round(latest.get(spread_key, 0) - prev_spread, 3)
    
    max_spread = max(spreads) if spreads else 0
    min_spread = min(spreads) if spreads else 0
    avg_spread = round(sum(spreads) / len(spreads), 3) if spreads else 0
    
    max_idx = spreads.index(max_spread) if spreads else -1
    min_idx = spreads.index(min_spread) if spreads else -1
    max_date = aligned_data[max_idx]["date"] if max_idx != -1 else "N/A"
    min_date = aligned_data[min_idx]["date"] if min_idx != -1 else "N/A"
    
    return {
        "latest_rate": latest.get(yield_key, 0),
        "latest_spread": latest.get(spread_key, 0),
        "spread_diff": spread_diff,
        "max_spread": max_spread,
        "max_date": max_date,
        "min_spread": min_spread,
        "min_date": min_date,
        "avg_spread": avg_spread
    }

def main():
    if not api_key:
        print("Error: BOK_API_KEY environment variable is not set in .env file.")
        return

    # 1. Calculate date range (last 5 years)
    end_dt = datetime.today()
    start_dt = end_dt - timedelta(days=5 * 365)
    
    start_date = start_dt.strftime("%Y%m%d")
    end_date = end_dt.strftime("%Y%m%d")
    
    print(f"Date Range: {start_date} ~ {end_date} (Last 5 Years)")
    
    # 2. Fetch all daily interest rate datasets
    print("Fetching Bank of Korea Base Interest Rate (기준금리)...")
    base_rate_rows = fetch_ecos_data("722Y001", "0101000", start_date, end_date)
    
    print("Fetching 3-Year Treasury Bond Yield (국고채 3년)...")
    treasury_rows = fetch_ecos_data("817Y002", "010200000", start_date, end_date)
    
    print("Fetching 3-Year Corporate Bond Yield (회사채 3년 AA-)...")
    corp_rows = fetch_ecos_data("817Y002", "010300000", start_date, end_date)
    
    print("Fetching 1-Year Industrial Finance Bond Yield (산금채 1년)...")
    ind_rows = fetch_ecos_data("817Y002", "010260000", start_date, end_date)
    
    print("Fetching 1-Year Monetary Stabilization Bond Yield (통안증권 1년)...")
    msb_rows = fetch_ecos_data("817Y002", "010400001", start_date, end_date)
    
    if not (base_rate_rows and treasury_rows and corp_rows and ind_rows and msb_rows):
        print("Failed to fetch required data series. Aborting.")
        return

    # 3. Create dictionaries for fast date-based lookup (TIME -> Value)
    base_rate_dict = {}
    for row in base_rate_rows:
        if row.get("DATA_VALUE"):
            try:
                base_rate_dict[row["TIME"]] = float(row["DATA_VALUE"])
            except ValueError:
                continue

    treasury_dict = {row["TIME"]: float(row["DATA_VALUE"]) for row in treasury_rows if row.get("DATA_VALUE")}
    corp_dict = {row["TIME"]: float(row["DATA_VALUE"]) for row in corp_rows if row.get("DATA_VALUE")}
    ind_dict = {row["TIME"]: float(row["DATA_VALUE"]) for row in ind_rows if row.get("DATA_VALUE")}
    msb_dict = {row["TIME"]: float(row["DATA_VALUE"]) for row in msb_rows if row.get("DATA_VALUE")}

    # 4. Align series on all trading business days (where Treasury 3Y is available)
    aligned_data = []
    for row in treasury_rows:
        date_str = row["TIME"]
        if date_str not in treasury_dict:
            continue
            
        tb_val = treasury_dict[date_str]
        br_val = get_rate_with_fallback(date_str, base_rate_dict)
        corp_val = get_rate_with_fallback(date_str, corp_dict)
        ind_val = get_rate_with_fallback(date_str, ind_dict)
        msb_val = get_rate_with_fallback(date_str, msb_dict)

        # Compute daily spreads
        spread_tb = round(tb_val - br_val, 3)
        spread_corp = round(corp_val - br_val, 3)
        spread_ind = round(ind_val - br_val, 3)
        spread_msb = round(msb_val - br_val, 3)
        
        # Format date as YYYY-MM-DD
        formatted_date = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}"
        
        aligned_data.append({
            "date": formatted_date,
            "base_rate": br_val,
            "treasury_3y": tb_val,
            "spread_tb": spread_tb,
            "corporate_3y": corp_val,
            "spread_corp": spread_corp,
            "industrial_1y": ind_val,
            "spread_ind": spread_ind,
            "stabilization_1y": msb_val,
            "spread_msb": spread_msb
        })

    # Sort chronologically (oldest to newest)
    aligned_data.sort(key=lambda x: x["date"])
    print(f"Successfully processed {len(aligned_data)} business days of aligned multi-spread data.")

    # 5. Write raw data to spread.csv
    csv_file = "spread.csv"
    with open(csv_file, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow([
            "Date", "Base_Rate", 
            "Treasury_3Y", "Spread_Treasury_3Y", 
            "Corporate_3Y", "Spread_Corporate_3Y", 
            "Industrial_1Y", "Spread_Industrial_1Y", 
            "Stabilization_1Y", "Spread_Stabilization_1Y"
        ])
        for d in aligned_data:
            writer.writerow([
                d["date"], d["base_rate"],
                d["treasury_3y"], d["spread_tb"],
                d["corporate_3y"], d["spread_corp"],
                d["industrial_1y"], d["spread_ind"],
                d["stabilization_1y"], d["spread_msb"]
            ])
    print(f"Successfully saved raw data to '{csv_file}'.")

    # 5b. Write raw data to spread.xlsx
    xlsx_file = "spread.xlsx"
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Spread Data"

        # Headers
        headers = [
            "Date", "Base_Rate", 
            "Treasury_3Y", "Spread_Treasury_3Y", 
            "Corporate_3Y", "Spread_Corporate_3Y", 
            "Industrial_1Y", "Spread_Industrial_1Y", 
            "Stabilization_1Y", "Spread_Stabilization_1Y"
        ]
        ws.append(headers)

        # Style Header
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Add Data
        for d in aligned_data:
            ws.append([
                d["date"], d["base_rate"],
                d["treasury_3y"], d["spread_tb"],
                d["corporate_3y"], d["spread_corp"],
                d["industrial_1y"], d["spread_ind"],
                d["stabilization_1y"], d["spread_msb"]
            ])

        # Style Data Rows
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        data_font = Font(name="Calibri", size=11)
        
        for row in range(2, len(aligned_data) + 2):
            cell_date = ws.cell(row=row, column=1)
            cell_date.alignment = Alignment(horizontal="center")
            
            for col in range(2, 11):
                cell = ws.cell(row=row, column=col)
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="right")

            for col in range(1, 11):
                cell = ws.cell(row=row, column=col)
                cell.font = data_font
                cell.border = thin_border

        # Auto-fit columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(xlsx_file)
        print(f"Successfully saved raw data to '{xlsx_file}'.")
    except Exception as ex:
        print(f"Error saving to Excel via openpyxl: {ex}")

    # 6. Calculate statistics for each bond series
    stats_tb = calculate_bond_stats(aligned_data, "treasury_3y", "spread_tb")
    stats_corp = calculate_bond_stats(aligned_data, "corporate_3y", "spread_corp")
    stats_ind = calculate_bond_stats(aligned_data, "industrial_1y", "spread_ind")
    stats_msb = calculate_bond_stats(aligned_data, "stabilization_1y", "spread_msb")
    
    # Calculate historical average divergence to find the largest anomaly
    tb_diff = stats_tb["latest_spread"] - stats_tb["avg_spread"]
    corp_diff = stats_corp["latest_spread"] - stats_corp["avg_spread"]
    ind_diff = stats_ind["latest_spread"] - stats_ind["avg_spread"]
    msb_diff = stats_msb["latest_spread"] - stats_msb["avg_spread"]
    
    diff_list = [
        {"name": "국고채 3년", "latest": stats_tb["latest_spread"], "avg": stats_tb["avg_spread"], "diff": tb_diff, "abs_diff": abs(tb_diff), "type": "Tb"},
        {"name": "회사채 3년 (AA-)", "latest": stats_corp["latest_spread"], "avg": stats_corp["avg_spread"], "diff": corp_diff, "abs_diff": abs(corp_diff), "type": "Corp"},
        {"name": "산금채 1년", "latest": stats_ind["latest_spread"], "avg": stats_ind["avg_spread"], "diff": ind_diff, "abs_diff": abs(ind_diff), "type": "Ind"},
        {"name": "통안증권 1년", "latest": stats_msb["latest_spread"], "avg": stats_msb["avg_spread"], "diff": msb_diff, "abs_diff": abs(msb_diff), "type": "Msb"}
    ]
    diff_list.sort(key=lambda x: x["abs_diff"], reverse=True)
    largest_anomaly = diff_list[0]
    
    latest_date = aligned_data[-1]["date"] if aligned_data else "N/A"
    
    stats = {
        "latest_date": latest_date,
        "tb": stats_tb,
        "corp": stats_corp,
        "ind": stats_ind,
        "msb": stats_msb,
        "anomaly": largest_anomaly
    }

    # 7. Generate beautiful, premium spread.html dashboard
    html_content = get_html_template(aligned_data, stats)
    
    html_file = "spread.html"
    with open(html_file, "w", encoding="utf-8") as f:
        f.write(html_content)
    print(f"Successfully created premium dashboard in '{html_file}'.")

    # Also synchronize to index.html
    index_file = "index.html"
    with open(index_file, "w", encoding="utf-8") as f:
        f.write(html_content)
    print(f"Successfully synchronized premium dashboard to '{index_file}'.")

def get_html_template(data_list, stats):
    data_json = json.dumps(data_list)
    stats_json = json.dumps(stats)

    return f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>한국은행 기준금리 대비 채권 스프레드 종합 대시보드</title>
    
    <!-- Premium Fonts -->
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=Outfit:wght@400;500;600;700;800&display=swap" rel="stylesheet">
    
    <!-- Chart.js via CDN -->
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    
    <!-- SheetJS via CDN -->
    <script src="https://cdn.jsdelivr.net/npm/xlsx@0.18.5/dist/xlsx.full.min.js"></script>
    
    <style>
        :root {{
            --bg-base: #080b11;
            --bg-surface: rgba(13, 18, 30, 0.75);
            --bg-card: rgba(22, 30, 49, 0.45);
            --border-glow: rgba(99, 102, 241, 0.15);
            --border-hover: rgba(99, 102, 241, 0.35);
            --text-primary: #f8fafc;
            --text-secondary: #94a3b8;
            --text-muted: #64748b;
            
            --color-base-rate: #a855f7; /* Neon Purple */
            --color-treasury: #3b82f6;  /* Neon Blue */
            --color-corporate: #f59e0b; /* Neon Amber */
            --color-industrial: #14b8a6;/* Neon Teal */
            --color-stabilization: #f43f5e; /* Neon Rose */
            --color-danger: #ef4444;    /* Crimson Red */
            
            --font-display: 'Outfit', sans-serif;
            --font-body: 'Inter', sans-serif;
        }}

        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}

        body {{
            background-color: var(--bg-base);
            background-image: 
                radial-gradient(circle at 15% 15%, rgba(99, 102, 241, 0.06) 0%, transparent 45%),
                radial-gradient(circle at 85% 85%, rgba(168, 85, 247, 0.06) 0%, transparent 50%);
            color: var(--text-primary);
            font-family: var(--font-body);
            min-height: 100vh;
            padding: 2rem 1.5rem;
            line-height: 1.5;
        }}

        .container {{
            max-width: 1400px;
            margin: 0 auto;
            display: flex;
            flex-direction: column;
            gap: 2rem;
        }}

        /* Header Styling */
        header {{
            display: flex;
            flex-direction: column;
            align-items: stretch;
            border-bottom: 1px solid var(--border-glow);
            padding-bottom: 1.5rem;
            gap: 1.25rem;
        }}

        .brand-row {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            gap: 1.5rem;
            flex-wrap: wrap;
        }}

        .brand-section h1 {{
            font-family: var(--font-display);
            font-size: 2.2rem;
            font-weight: 800;
            background: linear-gradient(135deg, #a855f7 0%, #3b82f6 30%, #f59e0b 65%, #14b8a6 100%);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            letter-spacing: -0.03em;
        }}

        .brand-section p {{
            font-size: 0.95rem;
            color: var(--text-secondary);
            margin-top: 0.25rem;
        }}

        .action-buttons {{
            display: flex;
            gap: 0.75rem;
        }}

        .btn {{
            background: rgba(255, 255, 255, 0.04);
            border: 1px solid var(--border-glow);
            color: var(--text-primary);
            padding: 0.6rem 1.2rem;
            border-radius: 8px;
            font-family: var(--font-body);
            font-size: 0.85rem;
            font-weight: 500;
            cursor: pointer;
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
            display: flex;
            align-items: center;
            gap: 0.5rem;
            backdrop-filter: blur(10px);
        }}

        .btn:hover {{
            background: rgba(99, 102, 241, 0.1);
            border-color: var(--border-hover);
            transform: translateY(-2px);
            box-shadow: 0 4px 12px rgba(99, 102, 241, 0.15);
        }}

        .btn-excel:hover {{
            background: rgba(16, 185, 129, 0.1) !important;
            border-color: rgba(16, 185, 129, 0.4) !important;
            color: #34d399 !important;
            transform: translateY(-2px);
            box-shadow: 0 4px 12px rgba(16, 185, 129, 0.15) !important;
        }}

        /* Anomaly Alert Banner */
        .anomaly-banner {{
            display: flex;
            align-items: center;
            gap: 0.75rem;
            background: rgba(245, 158, 11, 0.04);
            border: 1px solid rgba(245, 158, 11, 0.15);
            border-left: 4px solid var(--color-corporate);
            border-radius: 8px;
            padding: 0.65rem 1rem;
            font-size: 0.85rem;
            color: var(--text-primary);
            backdrop-filter: blur(12px);
            animation: pulse-glow 3.5s infinite ease-in-out;
            transition: all 0.3s ease;
        }}

        @keyframes pulse-glow {{
            0%, 100% {{ box-shadow: 0 0 5px rgba(245, 158, 11, 0.03); }}
            50% {{ box-shadow: 0 0 15px rgba(245, 158, 11, 0.12); }}
        }}

        .anomaly-icon {{
            font-size: 1.15rem;
            filter: drop-shadow(0 0 4px rgba(245, 158, 11, 0.4));
        }}

        .anomaly-text {{
            font-weight: 400;
            color: var(--text-secondary);
            width: 100%;
        }}

        .anomaly-text b {{
            font-weight: 700;
            color: var(--text-primary);
        }}

        /* Key Metrics Grid */
        .metrics-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
            gap: 1.5rem;
        }}

        .metric-card {{
            background: var(--bg-surface);
            backdrop-filter: blur(16px);
            border: 1px solid var(--border-glow);
            border-radius: 16px;
            padding: 1.5rem;
            display: flex;
            flex-direction: column;
            justify-content: space-between;
            position: relative;
            overflow: hidden;
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        }}

        .metric-card:hover {{
            border-color: var(--border-hover);
            transform: translateY(-3px);
            box-shadow: 0 8px 30px rgba(0, 0, 0, 0.4);
        }}

        .metric-card::before {{
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            width: 4px;
            height: 100%;
        }}

        .card-tb::before {{ background: var(--color-treasury); }}
        .card-corp::before {{ background: var(--color-corporate); }}
        .card-ind::before {{ background: var(--color-industrial); }}
        .card-msb::before {{ background: var(--color-stabilization); }}

        .metric-header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 0.5rem;
        }}

        .metric-label {{
            font-size: 0.85rem;
            color: var(--text-secondary);
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.05em;
        }}

        .metric-badge {{
            font-size: 0.72rem;
            padding: 0.15rem 0.4rem;
            border-radius: 5px;
            font-weight: 600;
            background: rgba(255, 255, 255, 0.06);
            color: var(--text-secondary);
        }}

        .metric-value-container {{
            display: flex;
            align-items: baseline;
            gap: 0.35rem;
            margin: 0.25rem 0 0.5rem 0;
        }}

        .metric-value {{
            font-family: var(--font-display);
            font-size: 2.1rem;
            font-weight: 700;
            letter-spacing: -0.02em;
        }}

        .metric-unit {{
            font-size: 0.95rem;
            color: var(--text-secondary);
            font-weight: 500;
        }}

        .metric-footer-grid {{
            display: flex;
            flex-direction: column;
            gap: 0.4rem;
            border-top: 1px solid rgba(255, 255, 255, 0.05);
            padding-top: 0.6rem;
            margin-top: 0.4rem;
        }}

        .footer-row {{
            display: flex;
            justify-content: space-between;
            font-size: 0.82rem;
            color: var(--text-secondary);
        }}

        .footer-row b {{
            color: var(--text-primary);
        }}

        .footer-row-secondary {{
            display: flex;
            justify-content: space-between;
            font-size: 0.75rem;
            color: var(--text-muted);
        }}

        .trend-up {{ color: var(--color-danger); font-weight: 600; }}
        .trend-down {{ color: var(--color-treasury); font-weight: 600; }}
        .trend-neutral {{ color: var(--text-muted); font-weight: 600; }}

        /* Dashboard Layout Grid */
        .dashboard-content {{
            display: grid;
            grid-template-columns: 2.5fr 1fr;
            gap: 1.5rem;
        }}

        @media (max-width: 1100px) {{
            .dashboard-content {{
                grid-template-columns: 1fr;
            }}
        }}

        .panel {{
            background: var(--bg-surface);
            backdrop-filter: blur(16px);
            border: 1px solid var(--border-glow);
            border-radius: 16px;
            padding: 1.75rem;
            display: flex;
            flex-direction: column;
            gap: 1.5rem;
            transition: border-color 0.3s ease;
        }}

        .panel:hover {{
            border-color: rgba(99, 102, 241, 0.25);
        }}

        .panel-title-bar {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            gap: 1rem;
            flex-wrap: wrap;
        }}

        .panel-title {{
            font-family: var(--font-display);
            font-size: 1.25rem;
            font-weight: 600;
            display: flex;
            align-items: center;
            gap: 0.6rem;
        }}

        .chart-view-selector {{
            display: flex;
            background: rgba(255, 255, 255, 0.03);
            border: 1px solid rgba(255, 255, 255, 0.05);
            padding: 0.2rem;
            border-radius: 8px;
        }}

        .view-btn {{
            background: transparent;
            border: none;
            color: var(--text-secondary);
            font-size: 0.78rem;
            font-weight: 600;
            padding: 0.35rem 0.9rem;
            border-radius: 6px;
            cursor: pointer;
            transition: all 0.2s ease;
        }}

        .view-btn.active {{
            background: rgba(99, 102, 241, 0.2);
            color: #818cf8;
            box-shadow: 0 2px 6px rgba(99, 102, 241, 0.1);
        }}

        .chart-controls {{
            display: flex;
            gap: 0.35rem;
            background: rgba(255, 255, 255, 0.03);
            padding: 0.2rem;
            border-radius: 8px;
            border: 1px solid rgba(255, 255, 255, 0.05);
        }}

        .chart-ctrl-btn {{
            background: transparent;
            border: none;
            color: var(--text-secondary);
            font-size: 0.75rem;
            font-weight: 600;
            padding: 0.35rem 0.75rem;
            border-radius: 6px;
            cursor: pointer;
            transition: all 0.2s ease;
        }}

        .chart-ctrl-btn.active {{
            background: rgba(99, 102, 241, 0.2);
            color: #818cf8;
        }}

        .chart-container {{
            position: relative;
            width: 100%;
            height: 420px;
        }}

        /* Legend Panel */
        .legend-list {{
            display: flex;
            flex-direction: column;
            gap: 0.85rem;
            background: rgba(255, 255, 255, 0.02);
            border-radius: 12px;
            padding: 1.1rem;
            border: 1px solid rgba(255, 255, 255, 0.04);
        }}

        .legend-item {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            font-size: 0.85rem;
            padding: 0.3rem 0;
        }}

        .legend-label-container {{
            display: flex;
            align-items: center;
            gap: 0.75rem;
            cursor: pointer;
            user-select: none;
            flex-grow: 1;
        }}

        .legend-color-dot {{
            width: 12px;
            height: 12px;
            border-radius: 3px;
        }}

        .legend-item.disabled {{
            opacity: 0.35;
            text-decoration: line-through;
        }}

        .legend-val {{
            font-weight: 600;
            font-family: var(--font-display);
            font-size: 0.9rem;
        }}

        /* Raw Data Grid Styling */
        .grid-panel {{
            grid-column: span 2;
        }}

        @media (max-width: 1100px) {{
            .grid-panel {{
                grid-column: span 1;
            }}
        }}

        .grid-toolbar {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            gap: 1.5rem;
            flex-wrap: wrap;
        }}

        .search-box {{
            position: relative;
            flex-grow: 1;
            max-width: 400px;
        }}

        .search-input {{
            width: 100%;
            background: rgba(255, 255, 255, 0.03);
            border: 1px solid var(--border-glow);
            color: var(--text-primary);
            padding: 0.6rem 1rem 0.6rem 2.2rem;
            border-radius: 8px;
            font-size: 0.85rem;
            outline: none;
            transition: all 0.3s ease;
        }}

        .search-input:focus {{
            border-color: var(--border-hover);
            background: rgba(255, 255, 255, 0.05);
            box-shadow: 0 0 10px rgba(99, 102, 241, 0.15);
        }}

        .search-icon {{
            position: absolute;
            left: 0.8rem;
            top: 50%;
            transform: translateY(-50%);
            width: 16px;
            height: 16px;
            fill: var(--text-secondary);
        }}

        /* Table Design */
        .table-wrapper {{
            width: 100%;
            overflow-x: auto;
            border-radius: 12px;
            border: 1px solid rgba(255, 255, 255, 0.05);
            background: rgba(255, 255, 255, 0.01);
        }}

        table {{
            width: 100%;
            border-collapse: collapse;
            text-align: left;
            font-size: 0.875rem;
            min-width: 1000px;
        }}

        th {{
            background: rgba(12, 17, 29, 0.65);
            color: var(--text-secondary);
            font-weight: 600;
            padding: 1.1rem 1rem;
            border-bottom: 1px solid rgba(255, 255, 255, 0.08);
            cursor: pointer;
            user-select: none;
            transition: all 0.2s ease;
        }}

        th:hover {{
            background: rgba(99, 102, 241, 0.1);
            color: var(--text-primary);
        }}

        th.sorted-asc::after {{ content: ' ▲'; font-size: 0.65rem; color: #818cf8; }}
        th.sorted-desc::after {{ content: ' ▼'; font-size: 0.65rem; color: #818cf8; }}

        td {{
            padding: 0.95rem 1rem;
            border-bottom: 1px solid rgba(255, 255, 255, 0.04);
            color: var(--text-secondary);
        }}

        tr:last-child td {{
            border-bottom: none;
        }}

        tr:hover td {{
            background: rgba(255, 255, 255, 0.015);
            color: var(--text-primary);
        }}

        .cell-date {{
            color: var(--text-primary);
            font-weight: 500;
        }}

        .cell-base-rate {{
            color: #c084fc;
            font-weight: 600;
        }}

        .cell-sub-spread {{
            font-size: 0.78rem;
            margin-top: 0.15rem;
            font-weight: 500;
        }}

        .cell-sub-spread.positive {{ color: #10b981; }}
        .cell-sub-spread.negative {{ color: #ef4444; }}

        /* Pagination Layout */
        .pagination-container {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-top: 1rem;
            font-size: 0.8rem;
            color: var(--text-secondary);
            flex-wrap: wrap;
            gap: 1rem;
        }}

        .pagination-buttons {{
            display: flex;
            gap: 0.35rem;
            align-items: center;
        }}

        .page-btn {{
            background: rgba(255, 255, 255, 0.03);
            border: 1px solid rgba(255, 255, 255, 0.05);
            color: var(--text-secondary);
            width: 32px;
            height: 32px;
            display: flex;
            justify-content: center;
            align-items: center;
            border-radius: 6px;
            cursor: pointer;
            transition: all 0.2s ease;
            font-weight: 500;
        }}

        .page-btn:hover:not(:disabled) {{
            background: rgba(99, 102, 241, 0.15);
            border-color: var(--border-hover);
            color: var(--text-primary);
        }}

        .page-btn.active {{
            background: rgba(99, 102, 241, 0.25);
            border-color: rgba(99, 102, 241, 0.4);
            color: #818cf8;
            font-weight: 600;
        }}

        .page-btn:disabled {{
            opacity: 0.3;
            cursor: not-allowed;
        }}
    </style>
</head>
<body>
    <div class="container">
        
        <!-- Header -->
        <header>
            <div class="brand-row">
                <div class="brand-section">
                    <h1>BOK SPREAD TERMINAL</h1>
                    <p>한국은행 기준금리 대비 주요 채권 스프레드 5개년 통합 분석 대시보드</p>
                </div>
                <div class="action-buttons">
                    <button class="btn" id="btnDownloadCsv">
                        <svg style="width:16px;height:16px;fill:currentColor" viewBox="0 0 24 24">
                            <path d="M5,20H19V18H5M19,9H15V3H9V9H5L12,16L19,9Z" />
                        </svg>
                        CSV 다운로드
                    </button>
                    <button class="btn btn-excel" id="btnDownloadExcel">
                        <svg style="width:16px;height:16px;fill:currentColor" viewBox="0 0 24 24">
                            <path d="M14 2H6C4.9 2 4 2.9 4 4V20C4 21.1 4.9 22 6 22H18C19.1 22 20 21.1 20 20V8L14 2M16 16H14V18H12V16H10V14H12V12H14V14H16V16M14 9V3.5L18.5 8H14V9Z" />
                        </svg>
                        엑셀 다운로드
                    </button>
                </div>
            </div>
            
            <!-- Anomaly Analysis Alert Banner -->
            <div class="anomaly-banner" id="anomalyBanner">
                <span class="anomaly-icon">📢</span>
                <span class="anomaly-text" id="anomalyText">스프레드 괴리율 분석 데이터 연동 중...</span>
            </div>
        </header>
        
        <!-- Metrics Grid -->
        <div class="metrics-grid">
            
            <!-- Treasury 3Y Spread -->
            <div class="metric-card card-tb">
                <div class="metric-header">
                    <span class="metric-label">국고채 3년 스프레드</span>
                    <span class="metric-badge">Gov Bond 3Y</span>
                </div>
                <div class="metric-value-container">
                    <span class="metric-value" id="valSpreadTb">0.00</span>
                    <span class="metric-unit">%p</span>
                </div>
                <div class="metric-footer-grid">
                    <div class="footer-row">
                        <span>최근 금리: <b id="valRateTb">0.00%</b></span>
                        <span id="diffSpreadTb" class="trend-neutral">0.00%p</span>
                    </div>
                    <div class="footer-row-secondary">
                        <span>평균: <span id="avgSpreadTb">0.00%p</span></span>
                        <span>범위: <span id="rangeSpreadTb">0.00 ~ 0.00</span></span>
                    </div>
                </div>
            </div>
            
            <!-- Corporate 3Y AA- Spread -->
            <div class="metric-card card-corp">
                <div class="metric-header">
                    <span class="metric-label">회사채 3년 스프레드</span>
                    <span class="metric-badge">Corp Bond 3Y</span>
                </div>
                <div class="metric-value-container">
                    <span class="metric-value" id="valSpreadCorp">0.00</span>
                    <span class="metric-unit">%p</span>
                </div>
                <div class="metric-footer-grid">
                    <div class="footer-row">
                        <span>최근 금리: <b id="valRateCorp">0.00%</b></span>
                        <span id="diffSpreadCorp" class="trend-neutral">0.00%p</span>
                    </div>
                    <div class="footer-row-secondary">
                        <span>평균: <span id="avgSpreadCorp">0.00%p</span></span>
                        <span>범위: <span id="rangeSpreadCorp">0.00 ~ 0.00</span></span>
                    </div>
                </div>
            </div>

            <!-- Industrial 1Y Spread -->
            <div class="metric-card card-ind">
                <div class="metric-header">
                    <span class="metric-label">산금채 1년 스프레드</span>
                    <span class="metric-badge">Ind Debenture 1Y</span>
                </div>
                <div class="metric-value-container">
                    <span class="metric-value" id="valSpreadInd">0.00</span>
                    <span class="metric-unit">%p</span>
                </div>
                <div class="metric-footer-grid">
                    <div class="footer-row">
                        <span>최근 금리: <b id="valRateInd">0.00%</b></span>
                        <span id="diffSpreadInd" class="trend-neutral">0.00%p</span>
                    </div>
                    <div class="footer-row-secondary">
                        <span>평균: <span id="avgSpreadInd">0.00%p</span></span>
                        <span>범위: <span id="rangeSpreadInd">0.00 ~ 0.00</span></span>
                    </div>
                </div>
            </div>

            <!-- MSB 1Y Spread -->
            <div class="metric-card card-msb">
                <div class="metric-header">
                    <span class="metric-label">통안증권 1년 스프레드</span>
                    <span class="metric-badge">MSB Bond 1Y</span>
                </div>
                <div class="metric-value-container">
                    <span class="metric-value" id="valSpreadMsb">0.00</span>
                    <span class="metric-unit">%p</span>
                </div>
                <div class="metric-footer-grid">
                    <div class="footer-row">
                        <span>최근 금리: <b id="valRateMsb">0.00%</b></span>
                        <span id="diffSpreadMsb" class="trend-neutral">0.00%p</span>
                    </div>
                    <div class="footer-row-secondary">
                        <span>평균: <span id="avgSpreadMsb">0.00%p</span></span>
                        <span>범위: <span id="rangeSpreadMsb">0.00 ~ 0.00</span></span>
                    </div>
                </div>
            </div>
            
        </div>
        
        <!-- Dashboard Content Grid -->
        <div class="dashboard-content">
            
            <!-- Interactive Graph Panel -->
            <div class="panel">
                <div class="panel-title-bar">
                    <div style="display:flex; align-items:center; gap:1rem;">
                        <h2 class="panel-title">
                            <svg style="width:18px;height:18px;fill:#818cf8" viewBox="0 0 24 24">
                                <path d="M16,11.78L20.24,4.45L21.97,5.45L16.74,14.5L10.23,10.75L5.46,19H22V21H2V3H4V17.54L9.5,8L16,11.78Z" />
                            </svg>
                            금리 및 스프레드 시계열 분석
                        </h2>
                        
                        <!-- Toggle view mode between yields and spreads -->
                        <div class="chart-view-selector">
                            <button class="view-btn active" id="btnViewYields" onclick="switchChartView('yields')">금리 추이</button>
                            <button class="view-btn" id="btnViewSpreads" onclick="switchChartView('spreads')">스프레드 추이</button>
                        </div>
                    </div>
                    
                    <div class="chart-controls">
                        <button class="chart-ctrl-btn" data-range="1y">1년</button>
                        <button class="chart-ctrl-btn" data-range="3y">3년</button>
                        <button class="chart-ctrl-btn active" data-range="5y">5년 전체</button>
                    </div>
                </div>
                
                <div class="chart-container">
                    <canvas id="spreadChart"></canvas>
                </div>
            </div>
            
            <!-- Side Panel Details -->
            <div class="panel">
                <h2 class="panel-title">
                    <svg style="width:18px;height:18px;fill:#c084fc" viewBox="0 0 24 24">
                        <path d="M11,9H13V7H11M12,20C7.59,20 4,16.41 4,12C4,7.59 7.59,4 12,4C16.41,4 20,7.59 20,12C20,16.41 16.41,20 12,20M12,2A10,10 0 0,0 2,12A10,10 0 0,0 12,22A10,10 0 0,0 22,12A10,10 0 0,0 12,2M11,17H13V11H11V17Z" />
                    </svg>
                    대시보드 범례 및 제어
                </h2>
                
                <div class="legend-list" id="legendContainer">
                    <!-- Injected dynamically by JS depending on view mode -->
                </div>

                <div style="font-size: 0.8rem; color: var(--text-secondary); display:flex; flex-direction:column; gap:0.55rem; border-top:1px solid rgba(255,255,255,0.05); padding-top:1rem;">
                    <p>💡 <b>스프레드(Spread)란?</b><br>각 채권 금리에서 기준금리를 뺀 수치입니다. 스프레드 수치는 발행 신용도 및 잔존 기간별 위험 프리미엄을 나타내는 지표입니다.</p>
                    <p>📊 <b>분석 대상 채권 안내:</b><br>
                        • <b>국고채 3년</b>: 정부 발행 무위험 벤치마크<br>
                        • <b>회사채 3년 (AA-)</b>: AAA급 우량 특수채/회사채 대표 지표<br>
                        • <b>산금채 1년</b>: 대표적인 우량 금융채 지표<br>
                        • <b>통안증권 1년</b>: 한은 발행 유동성 조절 중앙은행 채권
                    </p>
                </div>
            </div>
            
            <!-- Raw Data Grid Panel -->
            <div class="panel grid-panel">
                <div class="grid-toolbar">
                    <h2 class="panel-title" style="margin-right: auto">
                        <svg style="width:18px;height:18px;fill:#10b981" viewBox="0 0 24 24">
                            <path d="M10 4H4C2.9 4 2 4.9 2 6V18C2 19.1 2.9 20 4 20H20C21.1 20 22 19.1 22 18V8C22 6.9 21.1 6 20 6H12L10 4M4 8H20V18H4V8Z" />
                        </svg>
                        최근 5개년 Raw 데이터 그리드
                    </h2>
                    
                    <div class="search-box">
                        <svg class="search-icon" viewBox="0 0 24 24">
                            <path d="M9.5,3A6.5,6.5 0 0,1 16,9.5C16,11.11 15.41,12.59 14.44,13.73L14.71,14H15.5L20.5,19L19,20.5L14,15.5V14.71L13.73,14.44C12.59,15.41 11.11,16 9.5,16A6.5,6.5 0 0,1 3,9.5A6.5,6.5 0 0,1 9.5,3M9.5,5C7,5 5,7 5,9.5C5,12 7,14 9.5,14C12,14 14,12 14,9.5C14,7 12,5 9.5,5Z" />
                        </svg>
                        <input type="text" class="search-input" id="gridSearch" placeholder="날짜 또는 수치 검색...">
                    </div>
                </div>

                <div class="table-wrapper">
                    <table>
                        <thead>
                            <tr>
                                <th onclick="handleSort('date')" id="th-date" class="sorted-desc">조회 일자</th>
                                <th onclick="handleSort('base_rate')" id="th-base_rate">기준 금리 (%)</th>
                                <th onclick="handleSort('treasury_3y')" id="th-treasury_3y">국고채 3년 (%)</th>
                                <th onclick="handleSort('corporate_3y')" id="th-corporate_3y">회사채 3년 (%)</th>
                                <th onclick="handleSort('industrial_1y')" id="th-industrial_1y">산금채 1년 (%)</th>
                                <th onclick="handleSort('stabilization_1y')" id="th-stabilization_1y">통안증권 1년 (%)</th>
                            </tr>
                        </thead>
                        <tbody id="gridBody">
                            <!-- Injected by JS -->
                        </tbody>
                    </table>
                </div>

                <!-- Grid Pagination -->
                <div class="pagination-container">
                    <div class="pagination-info">
                        전체 <span id="lblTotalRecords" style="color:var(--text-primary); font-weight:600">0</span>건 중 
                        <span id="lblStartRecord">0</span> - <span id="lblEndRecord">0</span> 표시
                    </div>
                    <div class="pagination-buttons" id="paginationBtns">
                        <!-- Injected by JS -->
                    </div>
                </div>
            </div>

        </div>
    </div>

    <!-- Data Injection -->
    <script>
        const rawData = {data_json};
        const stats = {stats_json};
        
        let filteredData = [...rawData];
        let displayData = [...rawData];
        
        // Active display range and chart mode
        let activeRange = '5y';
        let activeChartMode = 'yields'; // 'yields' or 'spreads'
        let hiddenDatasets = {{
            'yields': [false, false, false, false, false],
            'spreads': [false, false, false, false]
        }};
        
        // Sorting and Pagination State
        let currentSortColumn = 'date';
        let currentSortDirection = 'desc'; // 'asc' or 'desc'
        let currentPage = 1;
        let rowsPerPage = 15;

        // Populate card values helper
        function setupMetricCard(prefix, statObj) {{
            document.getElementById(`valSpread${{prefix}}`).innerText = statObj.latest_spread.toFixed(2);
            document.getElementById(`valRate${{prefix}}`).innerText = `${{statObj.latest_rate.toFixed(2)}}%`;
            document.getElementById(`avgSpread${{prefix}}`).innerText = `${{statObj.avg_spread.toFixed(2)}}%p`;
            document.getElementById(`rangeSpread${{prefix}}`).innerText = `${{statObj.min_spread.toFixed(2)}} ~ +${{statObj.max_spread.toFixed(2)}}%p`;
            
            const diffEl = document.getElementById(`diffSpread${{prefix}}`);
            if (statObj.spread_diff > 0) {{
                diffEl.className = 'trend-up';
                diffEl.innerText = `+${{statObj.spread_diff.toFixed(3)}}%p ▲`;
            }} else if (statObj.spread_diff < 0) {{
                diffEl.className = 'trend-down';
                diffEl.innerText = `${{statObj.spread_diff.toFixed(3)}}%p ▼`;
            }} else {{
                diffEl.className = 'trend-neutral';
                diffEl.innerText = `0.000%p`;
            }}
        }}

        // Setup Metric Cards
        setupMetricCard('Tb', stats.tb);
        setupMetricCard('Corp', stats.corp);
        setupMetricCard('Ind', stats.ind);
        setupMetricCard('Msb', stats.msb);

        // Populate Dynamic Spread Anomaly Alert Banner
        const anomaly = stats.anomaly;
        const diffSign = anomaly.diff >= 0 ? '+' : '';
        const colorMap = {{
            'Tb': 'var(--color-treasury)',
            'Corp': 'var(--color-corporate)',
            'Ind': 'var(--color-industrial)',
            'Msb': 'var(--color-stabilization)'
        }};
        const activeColor = colorMap[anomaly.type] || 'var(--color-corporate)';
        
        // Set dynamic border-left and light background glow depending on active anomaly bond type
        const bannerEl = document.getElementById('anomalyBanner');
        bannerEl.style.borderLeftColor = activeColor;
        
        let glowColor = 'rgba(245, 158, 11, 0.04)'; // Default Corp
        let borderColor = 'rgba(245, 158, 11, 0.15)';
        if (anomaly.type === 'Tb') glowColor = 'rgba(59, 130, 246, 0.04)';
        if (anomaly.type === 'Tb') borderColor = 'rgba(59, 130, 246, 0.15)';
        if (anomaly.type === 'Ind') glowColor = 'rgba(20, 184, 166, 0.04)';
        if (anomaly.type === 'Ind') borderColor = 'rgba(20, 184, 166, 0.15)';
        if (anomaly.type === 'Msb') glowColor = 'rgba(244, 63, 94, 0.04)';
        if (anomaly.type === 'Msb') borderColor = 'rgba(244, 63, 94, 0.15)';
        bannerEl.style.backgroundColor = glowColor;
        bannerEl.style.borderColor = borderColor;
        
        // Calculate all bond spread divergences (latest_spread - avg_spread)
        const tbDivergence = stats.tb.latest_spread - stats.tb.avg_spread;
        const corpDivergence = stats.corp.latest_spread - stats.corp.avg_spread;
        const indDivergence = stats.ind.latest_spread - stats.ind.avg_spread;
        const msbDivergence = stats.msb.latest_spread - stats.msb.avg_spread;

        const formatDiv = (name, val) => {{
            const sign = val >= 0 ? '+' : '';
            const color = val >= 0 ? 'var(--color-danger)' : 'var(--color-treasury)';
            return `<span style="white-space: nowrap;">${{name}}: <b style="color: ${{color}}; font-weight: 600;">${{sign}}${{val.toFixed(2)}}%p</b></span>`;
        }};

        const allDivergencesText = [
            formatDiv('국고채 3년', tbDivergence),
            formatDiv('회사채 3년', corpDivergence),
            formatDiv('산금채 1년', indDivergence),
            formatDiv('통안증권 1년', msbDivergence)
        ].join(' <span style="color: var(--text-muted); margin: 0 0.5rem;">|</span> ');

        const highlightColor = anomaly.diff >= 0 ? 'var(--color-danger)' : 'var(--color-treasury)';

        document.getElementById('anomalyText').innerHTML = `
            <div style="display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap; width: 100%; gap: 1rem;">
                <div style="font-size: 0.85rem; color: var(--text-secondary); display: flex; align-items: center; gap: 0.25rem; flex-wrap: wrap;">
                    <span style="font-weight: 600; color: var(--text-muted); margin-right: 0.25rem;">[채권별 스프레드 괴리율]</span>
                    ${{allDivergencesText}}
                </div>
                <div style="font-size: 0.95rem; font-weight: 800; color: var(--text-primary); background: rgba(255,255,255,0.02); padding: 0.35rem 0.85rem; border-radius: 6px; border: 1px solid rgba(255,255,255,0.05); display: flex; align-items: center; gap: 0.35rem;">
                    🚨 <span style="font-weight: 500; color: var(--text-secondary);">최대 괴리:</span> 
                    <span style="color: ${{activeColor}}; font-size: 1.05rem; font-weight: 900; text-shadow: 0 0 10px ${{activeColor}}50;">${{anomaly.name}}</span>
                    <span style="font-size: 0.85rem; font-weight: 500; color: var(--text-muted); margin-left: 0.25rem;">(괴리폭: <b style="color: ${{highlightColor}}; font-weight: 800;">${{diffSign}}${{anomaly.diff.toFixed(2)}}%p</b>)</span>
                </div>
            </div>
        `;

        // --- Chart Configuration & Multi-Mode Logic ---
        const ctx = document.getElementById('spreadChart').getContext('2d');
        let myChart = null;

        // Custom legends builder
        function renderCustomLegends() {{
            const legendContainer = document.getElementById('legendContainer');
            legendContainer.innerHTML = '';
            
            if (activeChartMode === 'yields') {{
                const items = [
                    {{ name: '한국은행 기준금리 (BOK)', color: 'var(--color-base-rate)', val: '0.00%' }},
                    {{ name: '국고채 3년 금리', color: 'var(--color-treasury)', val: `${{stats.tb.latest_rate.toFixed(2)}}%` }},
                    {{ name: '회사채 3년 (AA-) 금리', color: 'var(--color-corporate)', val: `${{stats.corp.latest_rate.toFixed(2)}}%` }},
                    {{ name: '산금채 1년 금리', color: 'var(--color-industrial)', val: `${{stats.ind.latest_rate.toFixed(2)}}%` }},
                    {{ name: '통안증권 1년 금리', color: 'var(--color-stabilization)', val: `${{stats.msb.latest_rate.toFixed(2)}}%` }}
                ];
                
                if (rawData.length > 0) {{
                    items[0].val = `${{rawData[rawData.length - 1].base_rate.toFixed(2)}}%`;
                }}
                
                items.forEach((item, idx) => {{
                    const div = document.createElement('div');
                    div.className = `legend-item ${{hiddenDatasets.yields[idx] ? 'disabled' : ''}}`;
                    div.id = `legend-yields-${{idx}}`;
                    div.innerHTML = `
                        <div class="legend-label-container" onclick="toggleDataset(${{idx}})">
                            <div class="legend-color-dot" style="background:${{item.color}}"></div>
                            <span>${{item.name}}</span>
                        </div>
                        <span class="legend-val" style="color:${{item.color}}">${{item.val}}</span>
                    `;
                    legendContainer.appendChild(div);
                }});
            }} else {{
                const items = [
                    {{ name: '국고채 3년 스프레드', color: 'var(--color-treasury)', val: `${{stats.tb.latest_spread.toFixed(2)}}%p` }},
                    {{ name: '회사채 3년 스프레드', color: 'var(--color-corporate)', val: `${{stats.corp.latest_spread.toFixed(2)}}%p` }},
                    {{ name: '산금채 1년 스프레드', color: 'var(--color-industrial)', val: `${{stats.ind.latest_spread.toFixed(2)}}%p` }},
                    {{ name: '통안증권 1년 스프레드', color: 'var(--color-stabilization)', val: `${{stats.msb.latest_spread.toFixed(2)}}%p` }}
                ];
                
                items.forEach((item, idx) => {{
                    const div = document.createElement('div');
                    div.className = `legend-item ${{hiddenDatasets.spreads[idx] ? 'disabled' : ''}}`;
                    div.id = `legend-spreads-${{idx}}`;
                    div.innerHTML = `
                        <div class="legend-label-container" onclick="toggleDataset(${{idx}})">
                            <div class="legend-color-dot" style="background:${{item.color}}"></div>
                            <span>${{item.name}}</span>
                        </div>
                        <span class="legend-val" style="color:${{item.color}}">${{item.val}}</span>
                    `;
                    legendContainer.appendChild(div);
                }});
            }}
        }}

        // Switch Chart Views (Yields vs Spreads)
        function switchChartView(mode) {{
            if (activeChartMode === mode) return;
            activeChartMode = mode;
            
            document.getElementById('btnViewYields').classList.toggle('active', mode === 'yields');
            document.getElementById('btnViewSpreads').classList.toggle('active', mode === 'spreads');
            
            buildChart();
            renderCustomLegends();
        }}

        // Retrieve subset based on range (1y, 3y, 5y)
        function getRangeSubset() {{
            let sliceCount = rawData.length;
            if (activeRange === '1y') {{
                sliceCount = Math.min(rawData.length, 250);
            }} else if (activeRange === '3y') {{
                sliceCount = Math.min(rawData.length, 750);
            }}
            return rawData.slice(-sliceCount);
        }}

        // Build/Rebuild Chart.js instance
        function buildChart() {{
            if (myChart) {{
                myChart.destroy();
            }}
            
            const subset = getRangeSubset();
            const labels = subset.map(d => d.date);
            let datasets = [];
            
            if (activeChartMode === 'yields') {{
                datasets = [
                    {{
                        label: '기준금리 (%)',
                        data: subset.map(d => d.base_rate),
                        borderColor: '#a855f7',
                        borderWidth: 2,
                        pointRadius: 0,
                        pointHoverRadius: 5,
                        tension: 0.1,
                        hidden: hiddenDatasets.yields[0]
                    }},
                    {{
                        label: '국고채 3년 (%)',
                        data: subset.map(d => d.treasury_3y),
                        borderColor: '#3b82f6',
                        borderWidth: 2,
                        pointRadius: 0,
                        pointHoverRadius: 5,
                        tension: 0.1,
                        hidden: hiddenDatasets.yields[1]
                    }},
                    {{
                        label: '회사채 3년 AA- (%)',
                        data: subset.map(d => d.corporate_3y),
                        borderColor: '#f59e0b',
                        borderWidth: 2,
                        pointRadius: 0,
                        pointHoverRadius: 5,
                        tension: 0.1,
                        hidden: hiddenDatasets.yields[2]
                    }},
                    {{
                        label: '산금채 1년 (%)',
                        data: subset.map(d => d.industrial_1y),
                        borderColor: '#14b8a6',
                        borderWidth: 2,
                        pointRadius: 0,
                        pointHoverRadius: 5,
                        tension: 0.1,
                        hidden: hiddenDatasets.yields[3]
                    }},
                    {{
                        label: '통안증권 1년 (%)',
                        data: subset.map(d => d.stabilization_1y),
                        borderColor: '#f43f5e',
                        borderWidth: 2,
                        pointRadius: 0,
                        pointHoverRadius: 5,
                        tension: 0.1,
                        hidden: hiddenDatasets.yields[4]
                    }}
                ];
            }} else {{
                const gradTb = ctx.createLinearGradient(0, 0, 0, 400);
                gradTb.addColorStop(0, 'rgba(59, 130, 246, 0.08)');
                gradTb.addColorStop(1, 'rgba(59, 130, 246, 0.00)');

                const gradCorp = ctx.createLinearGradient(0, 0, 0, 400);
                gradCorp.addColorStop(0, 'rgba(245, 158, 11, 0.08)');
                gradCorp.addColorStop(1, 'rgba(245, 158, 11, 0.00)');
                
                datasets = [
                    {{
                        label: '국고채 3년 스프레드 (%p)',
                        data: subset.map(d => d.spread_tb),
                        borderColor: '#3b82f6',
                        borderWidth: 2.2,
                        pointRadius: 0,
                        pointHoverRadius: 5,
                        fill: true,
                        backgroundColor: gradTb,
                        tension: 0.1,
                        hidden: hiddenDatasets.spreads[0]
                    }},
                    {{
                        label: '회사채 3년 스프레드 (%p)',
                        data: subset.map(d => d.spread_corp),
                        borderColor: '#f59e0b',
                        borderWidth: 2.2,
                        pointRadius: 0,
                        pointHoverRadius: 5,
                        fill: true,
                        backgroundColor: gradCorp,
                        tension: 0.1,
                        hidden: hiddenDatasets.spreads[1]
                    }},
                    {{
                        label: '산금채 1년 스프레드 (%p)',
                        data: subset.map(d => d.spread_ind),
                        borderColor: '#14b8a6',
                        borderWidth: 2.2,
                        pointRadius: 0,
                        pointHoverRadius: 5,
                        tension: 0.1,
                        hidden: hiddenDatasets.spreads[2]
                    }},
                    {{
                        label: '통안증권 1년 스프레드 (%p)',
                        data: subset.map(d => d.spread_msb),
                        borderColor: '#f43f5e',
                        borderWidth: 2.2,
                        pointRadius: 0,
                        pointHoverRadius: 5,
                        tension: 0.1,
                        hidden: hiddenDatasets.spreads[3]
                    }}
                ];
            }}

            myChart = new Chart(ctx, {{
                type: 'line',
                data: {{
                    labels: labels,
                    datasets: datasets
                }},
                options: {{
                    responsive: true,
                    maintainAspectRatio: false,
                    interaction: {{
                        mode: 'index',
                        intersect: false
                    }},
                    plugins: {{
                        legend: {{
                            display: false
                        }},
                        tooltip: {{
                            backgroundColor: 'rgba(12, 17, 29, 0.95)',
                            titleFont: {{ family: 'Outfit', size: 13, weight: 'bold' }},
                            bodyFont: {{ family: 'Inter', size: 12 }},
                            borderColor: 'rgba(99, 102, 241, 0.25)',
                            borderWidth: 1,
                            padding: 12,
                            cornerRadius: 8,
                            callbacks: {{
                                label: function(context) {{
                                    let label = context.dataset.label || '';
                                    if (label) label += ': ';
                                    if (context.parsed.y !== null) {{
                                        label += context.parsed.y.toFixed(2);
                                        label += activeChartMode === 'yields' ? '%' : '%p';
                                    }}
                                    return label;
                                }}
                            }}
                        }}
                    }},
                    scales: {{
                        x: {{
                            grid: {{
                                color: 'rgba(255, 255, 255, 0.03)',
                                drawBorder: false
                            }},
                            ticks: {{
                                color: '#64748b',
                                font: {{ family: 'Inter', size: 11 }},
                                maxTicksLimit: 12
                            }}
                        }},
                        y: {{
                            grid: {{
                                color: 'rgba(255, 255, 255, 0.05)',
                                drawBorder: false
                            }},
                            ticks: {{
                                color: '#94a3b8',
                                font: {{ family: 'Inter', size: 11 }},
                                callback: function(value) {{
                                    return value.toFixed(1) + (activeChartMode === 'yields' ? '%' : '%p');
                                }}
                            }}
                        }}
                    }}
                }}
            }});
        }}

        // Handle legend items visibility toggles
        function toggleDataset(idx) {{
            let isCurrentlyHidden = false;
            
            if (activeChartMode === 'yields') {{
                hiddenDatasets.yields[idx] = !hiddenDatasets.yields[idx];
                isCurrentlyHidden = hiddenDatasets.yields[idx];
                document.getElementById(`legend-yields-${{idx}}`).classList.toggle('disabled', isCurrentlyHidden);
            }} else {{
                hiddenDatasets.spreads[idx] = !hiddenDatasets.spreads[idx];
                isCurrentlyHidden = hiddenDatasets.spreads[idx];
                document.getElementById(`legend-spreads-${{idx}}`).classList.toggle('disabled', isCurrentlyHidden);
            }}
            
            if (myChart) {{
                myChart.getDatasetMeta(idx).hidden = isCurrentlyHidden;
                myChart.data.datasets[idx].hidden = isCurrentlyHidden;
                myChart.update();
            }}
        }}

        // Range click handling
        document.querySelectorAll('.chart-ctrl-btn').forEach(btn => {{
            btn.addEventListener('click', (e) => {{
                document.querySelectorAll('.chart-ctrl-btn').forEach(b => b.classList.remove('active'));
                e.target.classList.add('active');
                
                activeRange = e.target.getAttribute('data-range');
                buildChart();
            }});
        }});

        // --- Grid Management ---

        // Render Raw Data Rows
        function renderTable() {{
            const tbody = document.getElementById('gridBody');
            tbody.innerHTML = '';
            
            if (displayData.length === 0) {{
                tbody.innerHTML = `<tr><td colspan="6" style="text-align: center; padding: 3rem; color: var(--text-muted)">검색 결과가 존재하지 않습니다.</td></tr>`;
                return;
            }}

            const startIdx = (currentPage - 1) * rowsPerPage;
            const endIdx = Math.min(startIdx + rowsPerPage, displayData.length);
            
            for (let i = startIdx; i < endIdx; i++) {{
                const row = displayData[i];
                const tr = document.createElement('tr');
                
                const fmtSpread = (val) => {{
                    const cls = val >= 0 ? 'positive' : 'negative';
                    const sign = val >= 0 ? '+' : '';
                    return `<div class="cell-sub-spread ${{cls}}">${{sign}}${{val.toFixed(2)}}%p</div>`;
                }};
                
                tr.innerHTML = `
                    <td class="cell-date">${{row.date}}</td>
                    <td class="cell-base-rate">${{row.base_rate.toFixed(2)}}%</td>
                    <td>
                        <div style="font-weight:600; color:var(--text-primary)">${{row.treasury_3y.toFixed(2)}}%</div>
                        ${{fmtSpread(row.spread_tb)}}
                    </td>
                    <td>
                        <div style="font-weight:600; color:var(--text-primary)">${{row.corporate_3y.toFixed(2)}}%</div>
                        ${{fmtSpread(row.spread_corp)}}
                    </td>
                    <td>
                        <div style="font-weight:600; color:var(--text-primary)">${{row.industrial_1y.toFixed(2)}}%</div>
                        ${{fmtSpread(row.spread_ind)}}
                    </td>
                    <td>
                        <div style="font-weight:600; color:var(--text-primary)">${{row.stabilization_1y.toFixed(2)}}%</div>
                        ${{fmtSpread(row.spread_msb)}}
                    </td>
                `;
                tbody.appendChild(tr);
            }}

            // Update Labels
            document.getElementById('lblTotalRecords').innerText = displayData.length;
            document.getElementById('lblStartRecord').innerText = displayData.length === 0 ? 0 : startIdx + 1;
            document.getElementById('lblEndRecord').innerText = endIdx;
            
            renderPaginationControls();
        }}

        // Render Pagination Button Controls
        function renderPaginationControls() {{
            const totalPages = Math.ceil(displayData.length / rowsPerPage);
            const container = document.getElementById('paginationBtns');
            container.innerHTML = '';

            if (totalPages <= 1) return;

            // Previous Button
            const prevBtn = document.createElement('button');
            prevBtn.className = 'page-btn';
            prevBtn.innerHTML = `&lt;`;
            prevBtn.disabled = currentPage === 1;
            prevBtn.onclick = () => {{
                if (currentPage > 1) {{
                    currentPage--;
                    renderTable();
                }}
            }};
            container.appendChild(prevBtn);

            let startPage = Math.max(1, currentPage - 2);
            let endPage = Math.min(totalPages, startPage + 4);
            if (endPage - startPage < 4) {{
                startPage = Math.max(1, endPage - 4);
            }}

            for (let p = startPage; p <= endPage; p++) {{
                const btn = document.createElement('button');
                btn.className = `page-btn ${{p === currentPage ? 'active' : ''}}`;
                btn.innerText = p;
                btn.onclick = () => {{
                    currentPage = p;
                    renderTable();
                }};
                container.appendChild(btn);
            }}

            // Next Button
            const nextBtn = document.createElement('button');
            nextBtn.className = 'page-btn';
            nextBtn.innerHTML = `&gt;`;
            nextBtn.disabled = currentPage === totalPages;
            nextBtn.onclick = () => {{
                if (currentPage < totalPages) {{
                    currentPage++;
                    renderTable();
                }}
            }};
            container.appendChild(nextBtn);
        }}

        // Handle Grid Sorting
        function handleSort(column) {{
            const thIds = ['th-date', 'th-base_rate', 'th-treasury_3y', 'th-corporate_3y', 'th-industrial_1y', 'th-stabilization_1y'];
            
            if (currentSortColumn === column) {{
                currentSortDirection = currentSortDirection === 'asc' ? 'desc' : 'asc';
            }} else {{
                currentSortColumn = column;
                currentSortDirection = 'desc';
            }}

            thIds.forEach(id => {{
                document.getElementById(id).className = '';
            }});

            document.getElementById(`th-${{column}}`).className = currentSortDirection === 'asc' ? 'sorted-asc' : 'sorted-desc';

            displayData.sort((a, b) => {{
                let valA = a[column];
                let valB = b[column];
                
                if (column === 'date') {{
                    valA = new Date(valA);
                    valB = new Date(valB);
                }}

                if (valA < valB) return currentSortDirection === 'asc' ? -1 : 1;
                if (valA > valB) return currentSortDirection === 'asc' ? 1 : -1;
                return 0;
            }});

            currentPage = 1;
            renderTable();
        }}

        // Real-Time Search Filtering
        const searchInput = document.getElementById('gridSearch');
        searchInput.addEventListener('input', (e) => {{
            const query = e.target.value.toLowerCase().trim();
            
            if (query === '') {{
                displayData = [...filteredData];
            }} else {{
                displayData = filteredData.filter(d => {{
                    return d.date.includes(query) || 
                           d.base_rate.toString().includes(query) || 
                           d.treasury_3y.toString().includes(query) || 
                           d.corporate_3y.toString().includes(query) || 
                           d.industrial_1y.toString().includes(query) || 
                           d.stabilization_1y.toString().includes(query);
                }});
            }}
            
            currentPage = 1;
            const savedCol = currentSortColumn;
            currentSortColumn = null;
            handleSort(savedCol);
        }});

        // CSV File Download Action
        document.getElementById('btnDownloadCsv').addEventListener('click', () => {{
            let csvContent = "data:text/csv;charset=utf-8,\\uFEFF";
            csvContent += "Date,Base_Rate,Treasury_3Y,Spread_Treasury_3Y,Corporate_3Y,Spread_Corporate_3Y,Industrial_1Y,Spread_Industrial_1Y,Stabilization_1Y,Spread_Stabilization_1Y\\n";
            
            rawData.forEach(row => {{
                csvContent += `${{row.date}},${{row.base_rate}},${{row.treasury_3y}},${{row.spread_tb}},${{row.corporate_3y}},${{row.spread_corp}},${{row.industrial_1y}},${{row.spread_ind}},${{row.stabilization_1y}},${{row.spread_msb}}\\n`;
            }});
            
            const encodedUri = encodeURI(csvContent);
            const link = document.createElement("a");
            link.setAttribute("href", encodedUri);
            link.setAttribute("download", "spread.csv");
            document.body.appendChild(link);
            link.click();
            document.body.removeChild(link);
        }});

        // Excel File Download Action
        document.getElementById('btnDownloadExcel').addEventListener('click', () => {{
            // Convert rawData to sheet with human-friendly headers
            const worksheetData = rawData.map(row => ({{
                "조회 일자": row.date,
                "기준 금리 (%)": row.base_rate,
                "국고채 3년 (%)": row.treasury_3y,
                "국고채 3년 스프레드 (%p)": row.spread_tb,
                "회사채 3년 (%)": row.corporate_3y,
                "회사채 3년 스프레드 (%p)": row.spread_corp,
                "산금채 1년 (%)": row.industrial_1y,
                "산금채 1년 스프레드 (%p)": row.spread_ind,
                "통안증권 1년 (%)": row.stabilization_1y,
                "통안증권 1년 스프레드 (%p)": row.spread_msb
            }}));

            const worksheet = XLSX.utils.json_to_sheet(worksheetData);
            
            // Adjust column widths for professional look
            const colWidths = [
                {{ wch: 14 }}, // Date
                {{ wch: 15 }}, // Base rate
                {{ wch: 15 }}, // Treasury
                {{ wch: 24 }}, // Treasury spread
                {{ wch: 15 }}, // Corp
                {{ wch: 24 }}, // Corp spread
                {{ wch: 15 }}, // Ind
                {{ wch: 24 }}, // Ind spread
                {{ wch: 15 }}, // stabilization
                {{ wch: 24 }}  // stabilization spread
            ];
            worksheet['!cols'] = colWidths;

            const workbook = XLSX.utils.book_new();
            XLSX.utils.book_append_sheet(workbook, worksheet, "5개년 Raw 데이터");
            
            XLSX.writeFile(workbook, "spread.xlsx");
        }});

        // Initial setup
        currentSortColumn = 'date';
        currentSortDirection = 'desc';
        displayData.sort((a, b) => new Date(b.date) - new Date(a.date));
        
        buildChart();
        renderCustomLegends();
        renderTable();
    </script>
</body>
</html>
"""

if __name__ == "__main__":
    main()
